"""Builds the .xlsx report: a 'Dashboard' sheet (KPIs + native Excel charts
for a quick visual read), a 'Latest comp set' sheet (side-by-side comparison
with colour-coded deltas), and a 'History' sheet (every run ever recorded,
for trend analysis in Excel / the Claude Excel add-in)."""
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RED_FONT = Font(color="B91C1C")    # negative % - red
GREEN_FONT = Font(color="0B7A3B")  # positive % - green

DASHBOARD_TITLE_FONT = Font(size=18, bold=True, color="1F2937")
DASHBOARD_SUBTITLE_FONT = Font(italic=True, color="6B7280")
DASHBOARD_NOTE_FONT = Font(italic=True, size=9, color="9CA3AF")
KPI_LABEL_FONT = Font(bold=True, size=9, color="6B7280")
KPI_VALUE_FONT = Font(size=11, bold=True, color="1F2937")
KPI_FILL = PatternFill("solid", fgColor="EFF6FF")

OWN_COLOR = "1F2937"       # our own price - dark slate, matches header colour
COMPETITOR_COLOR = "F97316"  # competitor price - orange


def _pct(v):
    return "" if v is None else f"{v:+.1f}%"


def _money(v):
    return "" if v is None else f"£{v:.2f}"


def _short_ts(ts: str) -> str:
    return ts[:16].replace("T", " ") if ts else ts


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _room_summary(comparison: list[dict]) -> list[dict]:
    """One row per St Mungo's room type: our current price alongside the
    min/avg/max of whichever competitor rooms are specified (via
    ROOM_EQUIVALENCE in compare.py) as equivalent to it. Feeds the
    Dashboard's bar chart."""
    own_rooms = {
        r["room_type"]: r for r in comparison
        if r["is_own"] and r.get("price_pw") is not None
    }
    comp_prices_by_own_room = defaultdict(list)
    for r in comparison:
        if not r["is_own"] and r.get("equivalent_room") and r.get("price_pw") is not None:
            comp_prices_by_own_room[r["equivalent_room"]].append(r["price_pw"])

    rows = []
    for room_type, r in own_rooms.items():
        prices = comp_prices_by_own_room.get(room_type, [])
        rows.append({
            "room_type": room_type,
            "category": r["category"],
            "our_price": r["price_pw"],
            "comp_min": min(prices) if prices else None,
            "comp_avg": (sum(prices) / len(prices)) if prices else None,
            "comp_max": max(prices) if prices else None,
            "comp_count": len(prices),
        })
    rows.sort(key=lambda x: (x["category"], x["our_price"]))
    return rows


def _trend_summary(history: list[dict]) -> list[dict]:
    """One row per run timestamp: the average own price and average
    competitor price that run, across every room with a recorded price.
    Feeds the Dashboard's trend line chart."""
    by_ts = defaultdict(lambda: {"own": [], "comp": []})
    for r in history:
        price = r.get("price_pw")
        if not price:
            continue
        bucket = by_ts[r["run_ts"]]
        bucket["own" if r.get("is_own") == "True" else "comp"].append(float(price))

    rows = []
    for ts in sorted(by_ts):
        b = by_ts[ts]
        rows.append({
            "run_ts": ts,
            "own_avg": (sum(b["own"]) / len(b["own"])) if b["own"] else None,
            "comp_avg": (sum(b["comp"]) / len(b["comp"])) if b["comp"] else None,
        })
    return rows


def _dashboard_kpis(comparison: list[dict]) -> dict:
    competitor_rooms = [r for r in comparison if not r["is_own"]]
    competitor_props = {r["property_name"] for r in competitor_rooms}
    matched = [r for r in competitor_rooms if r.get("vs_own_pct") is not None]
    avg_vs_own = (sum(r["vs_own_pct"] for r in matched) / len(matched)) if matched else None
    priced = [r for r in competitor_rooms if r.get("price_pw") is not None]
    return {
        "rooms_tracked": len(comparison),
        "competitors_tracked": len(competitor_props),
        "avg_vs_own": avg_vs_own,
        "cheapest": min(priced, key=lambda r: r["price_pw"], default=None),
        "priciest": max(priced, key=lambda r: r["price_pw"], default=None),
        "sold_out_count": len(competitor_rooms) - len(priced),
    }


def _build_dashboard(wb: Workbook, comparison: list[dict], history: list[dict]) -> None:
    ws = wb.create_sheet("Dashboard", 0)

    ws["A1"] = "St Mungo's Comp Set Dashboard"
    ws["A1"].font = DASHBOARD_TITLE_FONT
    ws.merge_cells("A1:F1")

    latest_ts = max((r["run_ts"] for r in comparison), default=None)
    ws["A2"] = f"Latest run: {_short_ts(latest_ts)}" if latest_ts else "No data yet"
    ws["A2"].font = DASHBOARD_SUBTITLE_FONT
    ws.merge_cells("A2:F2")

    kpis = _dashboard_kpis(comparison)

    def _room_label(r):
        return f"{r['property_name']} - {r['room_type']} ({_money(r['price_pw'])})" if r else "-"

    kpi_cards = [
        ("Rooms tracked", str(kpis["rooms_tracked"])),
        ("Competitors tracked", str(kpis["competitors_tracked"])),
        ("Avg competitor price vs us", _pct(kpis["avg_vs_own"]) or "-"),
        ("Cheapest competitor room", _room_label(kpis["cheapest"])),
        ("Priciest competitor room", _room_label(kpis["priciest"])),
        ("Sold-out competitor rooms", str(kpis["sold_out_count"])),
    ]
    for col, (label, value) in enumerate(kpi_cards, start=1):
        c1 = ws.cell(row=4, column=col, value=label)
        c1.font = KPI_LABEL_FONT
        c1.fill = KPI_FILL
        c1.alignment = Alignment(wrap_text=True, vertical="bottom")
        c2 = ws.cell(row=5, column=col, value=value)
        c2.font = KPI_VALUE_FONT
        c2.fill = KPI_FILL
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 30

    # --- Room summary table + bar chart -----------------------------------
    room_rows = _room_summary(comparison)
    room_header_row = 7
    ws.cell(row=room_header_row - 1, column=1,
            value="Room summary (source data for the chart to the right)").font = DASHBOARD_NOTE_FONT
    headers = ["Room", "Our price", "Competitor avg", "Competitor min", "Competitor max"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=room_header_row, column=col, value=h)
    _style_header(ws, row=room_header_row)
    for i, row in enumerate(room_rows, start=room_header_row + 1):
        ws.cell(row=i, column=1, value=row["room_type"])
        ws.cell(row=i, column=2, value=row["our_price"])
        ws.cell(row=i, column=3, value=row["comp_avg"])
        ws.cell(row=i, column=4, value=row["comp_min"])
        ws.cell(row=i, column=5, value=row["comp_max"])
    room_last_row = room_header_row + len(room_rows)

    if room_rows:
        bar = BarChart()
        bar.type = "col"
        bar.grouping = "clustered"
        bar.title = "Our price vs competitor price, by room"
        bar.y_axis.title = "£ per week"
        bar.x_axis.title = "Room"
        bar.height = 9
        bar.width = 24
        data = Reference(ws, min_col=2, max_col=3, min_row=room_header_row, max_row=room_last_row)
        cats = Reference(ws, min_col=1, min_row=room_header_row + 1, max_row=room_last_row)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.series[0].graphicalProperties.solidFill = OWN_COLOR
        bar.series[1].graphicalProperties.solidFill = COMPETITOR_COLOR
        ws.add_chart(bar, "G4")

    # --- Trend table + line chart -------------------------------------------
    trend_rows = _trend_summary(history)
    trend_header_row = room_last_row + 3
    ws.cell(row=trend_header_row - 1, column=1,
            value="Price trend over time (source data for the chart to the right)").font = DASHBOARD_NOTE_FONT
    trend_headers = ["Run", "Our avg price", "Competitor avg price"]
    for col, h in enumerate(trend_headers, start=1):
        ws.cell(row=trend_header_row, column=col, value=h)
    _style_header(ws, row=trend_header_row)
    for i, row in enumerate(trend_rows, start=trend_header_row + 1):
        ws.cell(row=i, column=1, value=_short_ts(row["run_ts"]))
        ws.cell(row=i, column=2, value=row["own_avg"])
        ws.cell(row=i, column=3, value=row["comp_avg"])
    trend_last_row = trend_header_row + len(trend_rows)

    if trend_rows:
        line = LineChart()
        line.title = "Average price trend over time"
        line.y_axis.title = "£ per week"
        line.x_axis.title = "Run"
        line.height = 9
        line.width = 24
        data = Reference(ws, min_col=2, max_col=3, min_row=trend_header_row, max_row=trend_last_row)
        cats = Reference(ws, min_col=1, min_row=trend_header_row + 1, max_row=trend_last_row)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.series[0].graphicalProperties.line.solidFill = OWN_COLOR
        line.series[0].graphicalProperties.line.width = 20000
        line.series[0].smooth = False
        line.series[1].graphicalProperties.line.solidFill = COMPETITOR_COLOR
        line.series[1].graphicalProperties.line.width = 20000
        line.series[1].smooth = False
        ws.add_chart(line, "G23")

    for col in ("A", "B", "C", "D", "E"):
        if ws.column_dimensions[col].width is None or ws.column_dimensions[col].width < 16:
            ws.column_dimensions[col].width = 16


def build_excel(comparison: list[dict], history: list[dict], path) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Latest comp set"
    headers = [
        "Property", "Room type", "Category", "Price (pw)", "Current offer",
        "vs last report", "vs baseline", "Equivalent St Mungo's room", "vs St Mungo's",
    ]
    ws.append(headers)
    _style_header(ws)

    ordered = sorted(comparison, key=lambda r: (not r["is_own"], r["property_name"], r["room_type"]))
    for r in ordered:
        ws.append([
            r["property_name"],
            r["room_type"],
            r["category"],
            _money(r["price_pw"]),
            r["offer_text"],
            _pct(r["pct_vs_prev"]),
            _pct(r["pct_vs_baseline"]),
            "-" if r["is_own"] else (r.get("equivalent_room") or ""),
            "-" if r["is_own"] else _pct(r["vs_own_pct"]),
        ])
        excel_row = ws.max_row
        for col, key in ((6, "pct_vs_prev"), (7, "pct_vs_baseline")):
            val = r.get(key)
            if val:
                ws.cell(row=excel_row, column=col).font = RED_FONT if val < 0 else GREEN_FONT

        # Same negative-red/positive-green convention as the trend columns
        # above - red = priced below us (bad for us), green = priced above
        # us (good for us).
        vs_own = r.get("vs_own_pct")
        if vs_own:
            ws.cell(row=excel_row, column=9).font = RED_FONT if vs_own < 0 else GREEN_FONT

    widths = [30, 32, 12, 12, 42, 16, 16, 30, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("History")
    if history:
        cols = list(history[0].keys())
        ws2.append(cols)
        _style_header(ws2)
        for r in history:
            ws2.append([r.get(c, "") for c in cols])
        for i in range(1, len(cols) + 1):
            ws2.column_dimensions[get_column_letter(i)].width = 20
        ws2.freeze_panes = "A2"

    _build_dashboard(wb, comparison, history)
    wb.active = 0  # Dashboard opens by default when the file is opened

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
